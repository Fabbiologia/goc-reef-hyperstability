#!/usr/bin/env python3
"""Ingest the raw CONAPESCA avisos de arribo files into one tidy parquet.

Input   data/CONAPESCA Raw/{2000..2017}.xlsx  and
        data/CONAPESCA Raw/AVISOS_*_{2018..2026}.csv
Output  data/conapesca_avisos_pacifico_2000_2026.csv.gz  (-> parquet in 01_data_preparation.R)
        data/conapesca_species_lookup.csv   (CLAVE ESPECIE -> NOMBRE CIENTIFICO)
        data/conapesca_ingest_log.csv       (rows read / kept per file)

These are the official government "avisos de arribo" (landing receipts): one
row per species per landing receipt. The three file generations differ in
column set, so every column is matched by a normalised header name, never by
position:

  2000-2007  33 cols, has NOMBRE CIENTIFICO, no CLAVE ESTADO
  2008-2017  36 cols, has NOMBRE CIENTIFICO + DESTINO CONSUMO
  2018-2026  35 cols (CSV, latin-1), NO NOMBRE CIENTIFICO, has permit fields

Because the CSV years carry no scientific name, we build a species lookup from
the xlsx years (CLAVE ESPECIE -> NOMBRE CIENTIFICO) and back-fill the CSV years
from it. This is what lets the analysis key on species rather than on the
common-name field.

We keep only Pacific-litoral wild-capture rows (LITORAL == PACIFICO and
ORIGEN PESCA == CAPTURA), which drops the Gulf of Mexico/Caribbean coast and
all aquaculture harvest. TIPO AVISO is kept as-is so the artisanal (MENORES)
and industrial (MAYORES) fleets can be separated downstream.

Run:  python3 scripts/01a_ingest_conapesca_raw.py
"""
import csv, gzip, os, re, sys, glob, unicodedata

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
RAW = os.path.join(ROOT, "data", "CONAPESCA Raw")
OUT = os.path.join(ROOT, "data", "conapesca_avisos_pacifico_2000_2026.csv.gz")
LOOKUP = os.path.join(ROOT, "data", "conapesca_species_lookup.csv")
LOG = os.path.join(ROOT, "data", "conapesca_ingest_log.csv")


def norm(h):
    """Header -> comparable key: strip accents/mojibake, keep A-Z0-9 only."""
    h = unicodedata.normalize("NFKD", str(h))
    h = "".join(c for c in h if not unicodedata.combining(c))
    return re.sub(r"[^A-Z0-9]", "", h.upper())


# canonical name -> the normalised header keys seen across the three generations
ALIASES = {
    "sitio_desembarque": ["NOMBRESITIODESEMBARQUE"],
    "estado":            ["NOMBREESTADO"],
    "clave_oficina":     ["CLAVEOFICINA"],
    "oficina":           ["NOMBREOFICINA"],
    "tipo_aviso":        ["TIPOAVISO"],
    "folio_aviso":       ["FOLIOAVISO"],
    "fecha_aviso":       ["FECHAAVISO"],
    "origen_pesca":      ["ORIGENPESCA"],  # NOT "ORIGEN": different column
    "lugar_captura":     ["NOMBRELUGARCAPTURA"],
    "n_embarcaciones":   ["NUMEROEMBARCACIONES"],
    "mes_corte":         ["MESCORTE"],
    "dias_efectivos":    ["DIASEFECTIVOS"],
    "duracion":          ["DURACION"],
    "nombre_principal":  ["NOMBREPRINCIPAL"],
    "clave_especie":     ["CLAVEESPECIE"],
    "nombre_especie":    ["NOMBREESPECIE"],
    "nombre_cientifico": ["NOMBRECIENTIFICO"],
    "peso_desembarcado_kg": ["PESODESEMBARCADOKILOGRAMOS"],
    "peso_vivo_kg":      ["PESOVIVOKILOGRAMOS"],
    "precio_pesos":      ["PRECIOPESOS"],
    "valor_pesos":       ["VALORPESOS"],
    "litoral":           ["LITORAL"],
}
# "AÑO CORTE" survives as ANOCORTE, AOCORTE (dropped enye) or AAOCORTE (mojibake)
YEAR_KEYS = {"ANOCORTE", "AOCORTE", "AAOCORTE", "ANIOCORTE"}

KEEP = list(ALIASES) + ["anio_corte", "mes_num", "fecha_iso", "source_year", "source_file"]

# MES CORTE is a Spanish month name and is the field the agency cuts the
# monthly statistics on, so it drives the time series (not FECHA AVISO, whose
# format differs between the xlsx and CSV generations).
MESES = {"ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
         "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11,
         "DICIEMBRE": 12}
NUMERIC = ("peso_desembarcado_kg", "peso_vivo_kg", "precio_pesos", "valor_pesos",
           "n_embarcaciones", "dias_efectivos", "duracion")


def build_index(header):
    """normalised header -> {canonical: column position}. Missing cols are fine."""
    pos = {}
    for i, h in enumerate(header):
        k = norm(h)
        if k in YEAR_KEYS:
            pos.setdefault("anio_corte", i)
            continue
        for canon, keys in ALIASES.items():
            if k in keys:
                pos.setdefault(canon, i)
    return pos


def num(v):
    """'1,629' / '' / None -> float. The CSV years quote thousands separators."""
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if not s or s in ("None", "NA", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s != "None" else None


def iso_date(v):
    """FECHA AVISO -> YYYY-MM-DD. xlsx gives '2000-08-04 00:00:00', CSV '02/01/2023'."""
    s = txt(v)
    if not s:
        return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return "%s-%s-%s" % m.groups()
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        return "%s-%02d-%02d" % (y, int(mo), int(d))
    return None


def emit(pos, row, year, fname, w, lookup, stats):
    """Filter one raw row and write it out. Streams: nothing is held in memory."""
    def g(c):
        i = pos.get(c)
        return row[i] if i is not None and i < len(row) else None

    lit = (txt(g("litoral")) or "").upper()
    org = (txt(g("origen_pesca")) or "").upper().strip()
    if lit != "PACIFICO":
        stats["drop_litoral"] += 1
        return
    # Aquaculture is dropped, but ORIGEN PESCA is blank on most rows of the
    # 2006-2017 files (those years also carry a separate ORIGEN column), so a
    # blank must be read as capture, not as "not capture". TIPO AVISO ==
    # COSECHA is the aquaculture-harvest receipt type and is the backstop.
    if org == "ACUACULTURA" or (txt(g("tipo_aviso")) or "").upper() == "COSECHA":
        stats["drop_acuacultura"] += 1
        return
    kg = num(g("peso_desembarcado_kg"))
    if kg is None or kg <= 0:
        stats["drop_sin_peso"] += 1
        return

    rec = {c: (num(g(c)) if c in NUMERIC else txt(g(c))) for c in ALIASES}
    y = num(g("anio_corte"))
    rec["anio_corte"] = int(y) if y else year
    rec["mes_num"] = MESES.get((rec["mes_corte"] or "").upper())
    rec["fecha_iso"] = iso_date(g("fecha_aviso"))
    rec["source_year"] = year
    rec["source_file"] = fname

    # scientific name: the xlsx years supply it and seed the lookup; the CSV
    # years have no such column, so they are back-filled by species code.
    code, sci = rec["clave_especie"], rec["nombre_cientifico"]
    if code and sci:
        lookup.setdefault(code, sci)
    elif code and not sci:
        rec["nombre_cientifico"] = lookup.get(code)
        if rec["nombre_cientifico"]:
            stats["sci_backfilled"] += 1
    if not rec["nombre_cientifico"]:
        stats["sin_cientifico"] += 1

    w.writerow([rec[c] for c in KEEP])
    stats["kept"] += 1


def read_xlsx(path, w, lookup, stats):
    import openpyxl
    year = int(os.path.splitext(os.path.basename(path))[0])
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    pos = None
    for row in ws.iter_rows(values_only=True):
        if pos is None:
            if norm(row[0] or "") == "RNPACTIVO":       # header row, after the banner
                pos = build_index(row)
            continue
        stats["read"] += 1
        emit(pos, row, year, os.path.basename(path), w, lookup, stats)
    wb.close()


def read_csv(path, w, lookup, stats):
    year = int(re.search(r"(\d{4})\.csv$", path).group(1))
    with open(path, encoding="latin-1", newline="") as fh:
        r = csv.reader(fh)
        pos = None
        for row in r:
            if pos is None:
                if row and norm(row[0]) == "RNPACTIVO":
                    pos = build_index(row)
                continue
            stats["read"] += 1
            emit(pos, row, year, os.path.basename(path), w, lookup, stats)


def main():
    files = sorted(glob.glob(os.path.join(RAW, "[12]*.xlsx"))) + \
            sorted(glob.glob(os.path.join(RAW, "AVISOS*.csv")))
    if not files:
        sys.exit("no raw CONAPESCA files found in %s" % RAW)

    lookup, log = {}, []
    with gzip.open(OUT, "wt", newline="", compresslevel=6) as fh:
        w = csv.writer(fh)
        w.writerow(KEEP)
        for p in files:
            stats = dict(read=0, kept=0, drop_litoral=0, drop_acuacultura=0,
                         drop_sin_peso=0, sci_backfilled=0, sin_cientifico=0)
            (read_xlsx if p.endswith(".xlsx") else read_csv)(p, w, lookup, stats)
            stats["file"] = os.path.basename(p)
            log.append(stats)
            print("  %-46s read %9d  kept %8d  no-sci %7d"
                  % (stats["file"], stats["read"], stats["kept"], stats["sin_cientifico"]),
                  flush=True)

    print("wrote %s (%.1f MB); species lookup: %d codes"
          % (OUT, os.path.getsize(OUT) / 1e6, len(lookup)))

    with open(LOOKUP, "w", newline="") as fh:
        w = csv.writer(fh); w.writerow(["clave_especie", "nombre_cientifico"])
        for k in sorted(lookup):
            w.writerow([k, lookup[k]])
    with open(LOG, "w", newline="") as fh:
        w = csv.DictWriter(fh, ["file", "read", "kept", "drop_litoral", "drop_acuacultura",
                                "drop_sin_peso", "sci_backfilled", "sin_cientifico"])
        w.writeheader(); w.writerows(log)
    print("wrote %s and %s" % (LOOKUP, LOG))


if __name__ == "__main__":
    main()
